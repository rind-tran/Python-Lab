from vnstock import *
import pandas
from openpyxl import load_workbook


MA_CK = "MWG"
OUTPUT_DIR = r"C:\Users\Adm\Desktop\CK\Danh Muc dau tu\ban le"
OUTPUT_FILE = f"{OUTPUT_DIR}\{MA_CK}.xlsx"
FREQUENCY = "Quarterly" # Quarterly | Yearly


def format_1000_separator(file_location):
    # Open excel to modify value
    wb = load_workbook(filename=file_location)
    for report in ("BalanceSheet", "IncomeStatement", "CashFlow"):
        # Get Sheet
        ws = wb[report]
        # Load source
        if report == "BalanceSheet":
            source = ws["B2":"K49"]
        elif report == "IncomeStatement":
            source = ws["B2":"K26"]
        else:
            source = ws["B2":"K52"]
        # Modify value
        for row in source:
            for cell in row:
                # cell.value = cell.value / 1000000
                cell.number_format = "#,##0"

    wb.save(OUTPUT_FILE)

# Export Balance Sheet
df = financial_report (symbol=MA_CK, report_type="BalanceSheet", frequency=FREQUENCY)
try:
    with pandas.ExcelWriter(OUTPUT_FILE, mode='a') as writer:
        df.to_excel(writer, index=False, sheet_name="BalanceSheet")
except:
    df.to_excel(OUTPUT_FILE, index=False, sheet_name="BalanceSheet")

# Export Income Statement
df = financial_report (symbol=MA_CK, report_type="IncomeStatement", frequency=FREQUENCY)
with pandas.ExcelWriter(OUTPUT_FILE, mode='a') as writer:
    df.to_excel(writer, index=False, sheet_name="IncomeStatement")

# Export Cash Flow
df = financial_report (symbol=MA_CK, report_type="CashFlow", frequency=FREQUENCY)
with pandas.ExcelWriter(OUTPUT_FILE, mode='a') as writer:
    df.to_excel(writer, index=False, sheet_name="CashFlow")

# Format separator
format_1000_separator(OUTPUT_FILE)
